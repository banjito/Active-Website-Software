"""
AMP Job Tracker — PROTOTYPE (2 jobs: 26011 multi-month, 26061 short).

Methodology (confirmed with Jeremy 2026-06-30):
 - Primary KPI = per-job GROSS MARGIN = (Revenue - fully-burdened direct cost) / Revenue, target 40%.
 - Revenue: actual billed (QBO invoices / job P&L income); quote as the plan.
 - Cost: per-job QBO P&L detail (authoritative, burdened, job-costed) by category;
   labor = max(P&L wage+burden accts, TimeActivity hrs x CostRate) so jobs whose
   wages weren't job-costed still get labor (e.g. 26061).
 - Hours: QBO TimeActivity, to the day -> time-phased curve.
 - Bid budget by category from the ampOS estimate (hours, mileage, per diem, etc.).
 - Overhead: company OpEx as % of revenue (~49% YTD) for a secondary net view.
"""
import json, collections, pathlib, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
import qbo_load as q

HERE = pathlib.Path(__file__).parent
DATA = HERE / "data"
OUT = HERE / "AMP Job Tracker - Prototype.xlsx"
TARGET = 0.40
PULL_DATE = "2026-06-30"

LABOR_ACCTS = {"5006", "5030", "6003", "6010"}  # wages + employer burden
ACCT_GROUP = {  # account prefix -> tracker category
    "5006": "Labor (W-2 + burden)", "5030": "Labor (W-2 + burden)",
    "6003": "Labor (W-2 + burden)", "6010": "Labor (W-2 + burden)",
    "5008": "Contract Labor", "5004": "Travel & Lodging", "5002": "Fuel/Mileage",
    "5001": "Materials", "5007": "Rental Equipment", "6390": "Per Diem",
    "6270": "Meals", "5003": "Meals", "5005": "Postage", "5009": "3rd-Party Testing",
    "5050": "Insurance",
}

# ---- load ampOS ----
jobs = {j["job_number"]: j for j in json.load(open(DATA / "jobs.json", encoding="utf-8"))}
opps = {o["id"]: o for o in json.load(open(DATA / "opportunities.json", encoding="utf-8"))}
ests = {}
for e in json.load(open(DATA / "estimates.json", encoding="utf-8")):
    ests.setdefault(e["opportunity_id"], []).append(e)

PROTO = [
    {"jn": "26011", "cust": "7028", "kind": "Multi-month", "start": "2026-01-30"},
    {"jn": "26061", "cust": "7235", "kind": "Short (multi-day)", "start": "2026-05-21"},
]

DONE_STATUS = {"billed", "ready_to_bill", "completed", "progress_billing"}


def eac(quoted, cost_to_date, billed, status):
    """Estimate-at-completion using BILLING % as the progress proxy.
    Projects final cost & margin instead of comparing cost-to-date to the full
    quote (which flatters early-stage jobs). Returns dict; proj_gm None when the
    job is mid-flight with no billing yet (can't project)."""
    quoted = float(quoted or 0)
    done = status in DONE_STATUS
    if done:
        pct = 1.0
    elif quoted and billed > 0:
        pct = min(billed / quoted, 1.0)
    else:
        pct = None  # in-progress, nothing billed -> cannot project
    target_cost = quoted * (1 - TARGET)          # cost ceiling to hit 40% GM
    if pct:
        proj_cost = cost_to_date / pct
        proj_gm = (quoted - proj_cost) / quoted if quoted else None
        expected_cost = target_cost * pct        # where cost SHOULD be at this % to hit 40%
        pace_var = cost_to_date - expected_cost   # +ve = over budget pace (margin eroding)
    else:
        proj_cost = proj_gm = expected_cost = pace_var = None
    return {"pct": pct, "proj_cost": proj_cost, "proj_gm": proj_gm,
            "target_cost": target_cost, "expected_cost": expected_cost, "pace_var": pace_var}


def bid_budget(jn):
    j = jobs[jn]; o = opps.get(j["opportunity_id"]) or {}
    elist = ests.get(j["opportunity_id"], [])
    e = sorted(elist, key=lambda x: x.get("created_at", ""))[-1] if elist else None
    d = json.loads(e["data"]) if e else {}
    hs = d.get("hoursSummary") or {}; td = d.get("travel_data") or {}
    sov = d.get("sovItems") or []
    mat = sum((it.get("quantity") or 0) * (it.get("materialPrice") or 0) for it in sov)
    mileage = sum(x.get("vehicleTravelCost", 0) for x in (td.get("travelExpense") or []) if isinstance(x, dict))
    travlabor = sum(x.get("totalTravelLabor", 0) for x in (td.get("travelTime") or []) if isinstance(x, dict))
    perdiem = sum((x.get("numDays", 0) or 0) * (x.get("firstDayRate", 0) or 0)
                  for x in (td.get("perDiem") or []) if isinstance(x, dict))
    return {
        "quote_no": o.get("quote_number"), "quoted": float(o.get("quoted_amount") or 0),
        "work_hrs": hs.get("workHours") or 0, "travel_hrs": hs.get("travelHours") or 0,
        "nonsov_hrs": hs.get("nonSovHours") or 0, "total_hrs": hs.get("totalHours") or 0,
        "men": hs.get("men"), "days": hs.get("daysOnsite"),
        "mileage": mileage, "travel_labor": travlabor, "per_diem": perdiem, "material": mat,
        "title": j.get("title"), "customer": o.get("amp_division"),
    }


def actuals(jn):
    pnl = q.load_qbo(DATA / f"qbo_pnl_{jn}.json")
    cols = [c["ColTitle"] for c in pnl["Columns"]["Column"]]; ai = cols.index("Amount")
    byacct = collections.defaultdict(float)

    def walk(rows, acct):
        for r in rows.get("Row", []):
            hdr = r.get("Header", {}).get("ColData"); a = acct
            if hdr and hdr[0].get("value"):
                a = hdr[0]["value"]
            if r.get("type") == "Data":
                cd = r.get("ColData"); amt = float(cd[ai]["value"]) if cd[ai].get("value") else 0
                byacct[a] += amt
            if "Rows" in r:
                walk(r["Rows"], a)
    walk(pnl["Rows"], "(root)")
    income = sum(v for a, v in byacct.items() if a[:1] == "4")
    labor_pnl = sum(v for a, v in byacct.items() if a[:4] in LABOR_ACCTS)
    cats = collections.defaultdict(float)
    for a, v in byacct.items():
        if a[:1] == "4":
            continue
        cats[ACCT_GROUP.get(a[:4], "Other")] += v
    t = q.time_rollup(DATA / f"qbo_time_{jn}.json")
    inv = q.invoice_rollup(DATA / f"qbo_inv_{jn}.json")
    # labor: fill from TimeActivity if P&L didn't job-cost wages
    labor = max(labor_pnl, t["cost"])
    if labor_pnl < t["cost"]:
        cats["Labor (W-2 + burden)"] = labor  # replace with TimeActivity-based labor
    total_cost = sum(cats.values())
    return {"income": income, "billed": inv["billed"], "collected": inv["collected"],
            "open": inv["open"], "labor_pnl": labor_pnl, "labor": labor,
            "hours": t["hours"], "by_day": t["by_day"], "cats": dict(cats),
            "total_cost": total_cost}


# ---------- styles ----------
H1 = Font(bold=True, size=15, color="1F3864")
H2 = Font(bold=True, size=12, color="1F3864")
HEADF = Font(bold=True, color="FFFFFF", size=10)
HEADFILL = PatternFill("solid", fgColor="1F3864")
SUB = Font(italic=True, size=9, color="666666")
BOLD = Font(bold=True)
GREEN = PatternFill("solid", fgColor="C6EFCE"); REDF = PatternFill("solid", fgColor="FFC7CE")
YEL = PatternFill("solid", fgColor="FFEB9C")
thin = Side(style="thin", color="BFBFBF"); BORD = Border(thin, thin, thin, thin)
MONEY = '#,##0;[Red](#,##0)'; PCT = '0.0%'; NUM1 = '#,##0.0'; NUM0 = '#,##0'

wb = openpyxl.Workbook()


def hdr_row(ws, row, labels, start=1):
    for i, l in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=l)
        c.fill = HEADFILL; c.font = HEADF
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        c.border = BORD


# ===== Company Margin tab =====
ws = wb.active; ws.title = "Company Margin"; ws.sheet_view.showGridLines = False
ws["A1"] = "AMP Quality Energy Services — Company Margin (YTD Jan–Jun 2026)"; ws["A1"].font = H1
ws["A2"] = f"From QBO P&L, accrual. The 40% target is a per-job GROSS margin floor; this shows where the company sits overall."; ws["A2"].font = SUB
waterfall = [
    ("Revenue", 3398268, 1.0, False),
    ("Direct job cost (COGS)", -1362049, -0.401, False),
    ("Gross Profit", 2036220, 0.599, True),
    ("Overhead / Operating expense", -1679004, -0.494, False),
    ("Net Operating Income", 357215, 0.105, True),
    ("Other income / adj (net)", 142468, 0.042, False),
    ("Net Income", 499683, 0.147, True),
]
hdr_row(ws, 4, ["Line", "Amount $", "% of Revenue"])
r = 5
for lab, amt, pct, sub in waterfall:
    ws.cell(row=r, column=1, value=lab).font = BOLD if sub else Font()
    ws.cell(row=r, column=2, value=amt).number_format = MONEY
    pc = ws.cell(row=r, column=3, value=pct); pc.number_format = PCT
    if sub:
        for cc in range(1, 4):
            ws.cell(row=r, column=cc).fill = PatternFill("solid", fgColor="D9E1F2")
    r += 1
ws.cell(row=r + 1, column=1, value="Gross margin 60% (well above the 40% job floor); net operating only 10.5% — overhead is the drag. The job tracker flags individual jobs that fall below 40% gross.").font = SUB
ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=6)
for w, i in zip([34, 16, 14], [1, 2, 3]):
    ws.column_dimensions[get_column_letter(i)].width = w


# ===== Job Scorecard tab =====
ws = wb.create_sheet("Job Scorecard"); ws.sheet_view.showGridLines = False
ws["A1"] = "Job Scorecard — Are we ahead or behind the quote?"; ws["A1"].font = H1
ws["A2"] = f"As of {PULL_DATE}. Gross margin vs 40% target. Green ≥40%, red <40%."; ws["A2"].font = SUB
metrics = [
    ("Job #", "jn"), ("Title", "title"), ("Type", "kind"), ("Status", "status"),
    ("Quoted $", "quoted"), ("Billed $", "billed"), ("Collected $", "collected"), ("Open AR $", "open"),
    ("— BUDGET vs ACTUAL —", None),
    ("Budget hours", "bud_hrs"), ("Actual hours", "act_hrs"), ("Hours variance %", "hrs_var"),
    ("— COST —", None),
    ("Labor (burdened) $", "labor"), ("Other direct $", "other"), ("Cost to date $", "total_cost"),
    ("— PROGRESS & PROJECTED MARGIN —", None),
    ("% complete (billed ÷ quote)", "pct"),
    ("Projected total cost (EAC) $", "proj_cost"),
    ("Projected gross margin % (at completion)", "proj_gm"),
    ("vs 40% target (projected)", "flag"),
    ("— pace check —", None),
    ("Cost budget for 40% GM $", "target_cost"),
    ("Expected cost @ % complete $", "expected_cost"),
    ("Cost-pace variance $ (actual − expected)", "pace_var"),
    ("GM % on billed-to-date (snapshot)", "gmpct_billed"),
]
# build per-job dicts
cards = []
for p in PROTO:
    jn = p["jn"]; b = bid_budget(jn); a = actuals(jn)
    rev = a["billed"] or a["income"]
    other = a["total_cost"] - a["labor"]
    e = eac(b["quoted"], a["total_cost"], a["billed"], jobs[jn].get("status"))
    card = {
        "jn": jn, "title": b["title"], "kind": p["kind"], "status": jobs[jn].get("status"),
        "quoted": b["quoted"], "billed": a["billed"], "collected": a["collected"], "open": a["open"],
        "bud_hrs": b["total_hrs"], "act_hrs": a["hours"],
        "hrs_var": (a["hours"] - b["total_hrs"]) / b["total_hrs"] if b["total_hrs"] else None,
        "labor": a["labor"], "other": other, "total_cost": a["total_cost"],
        "pct": e["pct"], "proj_cost": e["proj_cost"], "proj_gm": e["proj_gm"],
        "target_cost": e["target_cost"], "expected_cost": e["expected_cost"], "pace_var": e["pace_var"],
        "gmpct_billed": ((rev - a["total_cost"]) / rev) if rev else None,
    }
    if e["proj_gm"] is None:
        card["flag"] = "no billing yet"
    else:
        card["flag"] = "ON TARGET" if e["proj_gm"] >= TARGET else "BELOW 40%"
    cards.append(card)
# write transposed: metrics down rows, jobs across cols
hdr_row(ws, 4, ["Metric"] + [f'{c["jn"]}' for c in cards])
r = 5
for label, key in metrics:
    cell = ws.cell(row=r, column=1, value=label)
    if key is None:
        cell.font = H2
        for ci in range(2, 2 + len(cards)):
            ws.cell(row=r, column=ci).fill = PatternFill("solid", fgColor="EDEDED")
        r += 1; continue
    for ci, c in enumerate(cards, start=2):
        v = c.get(key)
        cc = ws.cell(row=r, column=ci, value=v)
        if key in ("quoted", "billed", "collected", "open", "labor", "other", "total_cost",
                   "proj_cost", "target_cost", "expected_cost", "pace_var"):
            cc.number_format = MONEY
        elif key in ("hrs_var", "gmpct_billed", "pct", "proj_gm"):
            cc.number_format = PCT
        elif key in ("bud_hrs", "act_hrs"):
            cc.number_format = NUM1
        if key == "proj_gm" and isinstance(v, (int, float)):
            cc.fill = GREEN if v >= TARGET else REDF; cc.font = BOLD
        if key == "flag":
            cc.fill = {"ON TARGET": GREEN, "BELOW 40%": REDF}.get(v, YEL); cc.font = BOLD
        if key == "pace_var" and isinstance(v, (int, float)):
            cc.fill = REDF if v > 0 else GREEN  # over expected burn = margin eroding
        if key == "hrs_var" and isinstance(v, (int, float)) and v > 0.10:
            cc.fill = YEL
    r += 1
ws.column_dimensions["A"].width = 30
for ci in range(2, 2 + len(cards)):
    ws.column_dimensions[get_column_letter(ci)].width = 22
ws.freeze_panes = "B5"


# ===== per-job detail tabs with daily curve =====
def detail_tab(p):
    jn = p["jn"]; b = bid_budget(jn); a = actuals(jn)
    ws = wb.create_sheet(f"{jn} Detail"); ws.sheet_view.showGridLines = False
    ws["A1"] = f'{jn} — {b["title"]}'; ws["A1"].font = H1
    ws["A2"] = f'{p["kind"]} · quote #{b["quote_no"]} · status {jobs[jn].get("status")} · as of {PULL_DATE}'; ws["A2"].font = SUB
    # category bid vs actual
    ws["A4"] = "Cost by category — Bid budget vs Actual"; ws["A4"].font = H2
    hdr_row(ws, 5, ["Category", "Bid $ (where quoted)", "Actual $", "Notes"])
    cat = a["cats"]
    rows = [
        ("Labor (W-2 + burden)", "", cat.get("Labor (W-2 + burden)", 0), f'{a["hours"]:.0f} actual hrs (budget {b["total_hrs"]:.0f} hrs)'),
        ("Contract Labor", "", cat.get("Contract Labor", 0), ""),
        ("Travel & Lodging", b["travel_labor"] + b["mileage"], cat.get("Travel & Lodging", 0) + cat.get("Fuel/Mileage", 0), "bid = travel labor + mileage"),
        ("Per Diem / Meals", b["per_diem"], cat.get("Per Diem", 0) + cat.get("Meals", 0), ""),
        ("Materials", b["material"], cat.get("Materials", 0), ""),
        ("Rental Equipment", "", cat.get("Rental Equipment", 0), ""),
        ("Other", "", cat.get("Other", 0) + cat.get("3rd-Party Testing", 0) + cat.get("Postage", 0) + cat.get("Insurance", 0), ""),
    ]
    rr = 6
    for nm, bidv, actv, note in rows:
        ws.cell(row=rr, column=1, value=nm)
        bc = ws.cell(row=rr, column=2, value=bidv if bidv != "" else None); bc.number_format = MONEY
        ac = ws.cell(row=rr, column=3, value=actv); ac.number_format = MONEY
        ws.cell(row=rr, column=4, value=note).font = SUB
        rr += 1
    ws.cell(row=rr, column=1, value="TOTAL COST").font = BOLD
    ws.cell(row=rr, column=3, value=f"=SUM(C6:C{rr-1})").number_format = MONEY
    ws.cell(row=rr, column=3).font = BOLD
    # margin summary
    rev = a["billed"] or a["income"]
    e = eac(b["quoted"], a["total_cost"], a["billed"], jobs[jn].get("status"))
    mr = rr + 2
    ws.cell(row=mr, column=1, value="MARGIN — projected at completion (not flattered by stage)").font = H2
    summ = [
        ("Quoted (plan)", b["quoted"], MONEY, False),
        ("Billed revenue", a["billed"], MONEY, False),
        ("Collected", a["collected"], MONEY, False),
        ("Cost to date", a["total_cost"], MONEY, False),
        ("% complete (billed ÷ quote)", e["pct"], PCT, False),
        ("Projected total cost (EAC)", e["proj_cost"], MONEY, False),
        ("Projected gross margin % (at completion)", e["proj_gm"], PCT, True),
        ("Target", TARGET, PCT, False),
        ("Cost-pace variance (actual − expected @ %)", e["pace_var"], MONEY, False),
        ("GM % on billed-to-date (snapshot only)", ((rev - a["total_cost"]) / rev) if rev else None, PCT, False),
    ]
    for i, (lab, val, fmt, isgm) in enumerate(summ):
        ws.cell(row=mr + 1 + i, column=1, value=lab)
        c = ws.cell(row=mr + 1 + i, column=2, value=val); c.number_format = fmt
        if isgm and isinstance(val, (int, float)):
            c.fill = GREEN if val >= TARGET else REDF; c.font = BOLD
        if lab.startswith("Cost-pace") and isinstance(val, (int, float)):
            c.fill = REDF if val > 0 else GREEN
    # daily cumulative curve table (for chart)
    cstart = 6
    ccol = 7  # column G
    ws.cell(row=5, column=ccol, value="Date"); ws.cell(row=5, column=ccol + 1, value="Cum. Hours"); ws.cell(row=5, column=ccol + 2, value="Cum. Labor $")
    hdr_row(ws, 5, ["Date", "Cum Hours", "Cum Labor $"], start=ccol)
    days = sorted(a["by_day"].items())
    cumh = cumc = 0; ri = 6
    for d, (h, c) in days:
        cumh += h; cumc += c
        ws.cell(row=ri, column=ccol, value=d)
        ws.cell(row=ri, column=ccol + 1, value=round(cumh, 1)).number_format = NUM1
        ws.cell(row=ri, column=ccol + 2, value=round(cumc)).number_format = MONEY
        ri += 1
    # chart
    if ri > 7:
        ch = LineChart(); ch.title = f"{jn} cumulative labor hours (to-the-day)"
        ch.height = 8; ch.width = 18
        data = Reference(ws, min_col=ccol + 1, min_row=5, max_row=ri - 1)
        cats = Reference(ws, min_col=ccol, min_row=6, max_row=ri - 1)
        ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
        ws.add_chart(ch, f"A{mr + 11}")
    for col, w in zip("ABCD", [26, 18, 16, 34]):
        ws.column_dimensions[col].width = w
    ws.column_dimensions[get_column_letter(ccol)].width = 12
    ws.column_dimensions[get_column_letter(ccol + 1)].width = 12
    ws.column_dimensions[get_column_letter(ccol + 2)].width = 14


for p in PROTO:
    detail_tab(p)

wb.save(OUT)
print("Saved:", OUT)
print("Tabs:", wb.sheetnames)
