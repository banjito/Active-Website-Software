"""
Build the AMP Job Profitability Excel model from the ampOS data pulled into ./data.

Bid side (authoritative) = ampOS. Actual side = QuickBooks (AMPQES), filled later.
Headline bid value = opportunities.quoted_amount (the estimate JSON's stored
grandTotal/calculatedValues are unreliable; only line items + labor HOURS are used
from the estimate). See memory: ampos-data-access.
"""
import json, collections, pathlib, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

HERE = pathlib.Path(__file__).parent
DATA = HERE / "data"
OUT = HERE / "AMP Job Profitability Model.xlsx"
PULL_DATE = "2026-06-30"

# ---------- load ----------
opps = json.load(open(DATA / "opportunities.json", encoding="utf-8"))
ests = json.load(open(DATA / "estimates.json", encoding="utf-8"))
jobs = json.load(open(DATA / "jobs.json", encoding="utf-8"))
custs = {c["id"]: c for c in json.load(open(DATA / "customers.json", encoding="utf-8"))}

# ---------- helpers ----------
DIV_MAP = {
    "alabama": "Alabama", "north_alabama": "Alabama", "Alabama": "Alabama",
    "tennessee": "Tennessee", "Tennessee": "Tennessee",
    "georgia": "Georgia", "Georgia": "Georgia",
    "international": "International", "Government/International": "International",
    "engineering": "Engineering", "Engineering": "Engineering",
    "scavenger": "Scavenger",
}
def norm_div(d):
    if not d:
        return "(unset)"
    return DIV_MAP.get(d, DIV_MAP.get(str(d).lower(), str(d)))

STATUS_LABEL = {
    "awareness": "Awareness", "interest": "Interest", "quote": "Quote",
    "decision": "Decision (open)", "decision - forecasted win": "Decision-Fcst Win",
    "awarded": "Awarded", "lost": "Lost", "no quote": "No Quote",
}
TYPE_LABEL = {
    "large_acceptance": "Large Acceptance", "small_acceptance": "Small Acceptance",
    "maintenance": "Maintenance", "engineering": "Engineering",
    "time_materials": "Time & Materials", "other": "Other",
}
def cust_name(cid):
    c = custs.get(cid)
    if not c:
        return ""
    return c.get("company_name") or c.get("name") or ""

def num(v):
    try:
        return float(v or 0)
    except Exception:
        return 0.0

# latest estimate per opportunity + extracted breakdown
est_by_opp = {}
for e in ests:
    est_by_opp.setdefault(e["opportunity_id"], []).append(e)

def latest_est(oid):
    lst = est_by_opp.get(oid)
    if not lst:
        return None
    return sorted(lst, key=lambda x: x.get("created_at", ""))[-1]

def est_breakdown(e):
    """Return dict: n_items, material$, expense$, labor_hours, men, sovItems[]"""
    try:
        d = json.loads(e["data"])
    except Exception:
        return None
    sov = d.get("sovItems") or []
    cv = d.get("calculatedValues") or {}
    hs = d.get("hoursSummary") or {}
    mat = sum(num(it.get("quantity")) * num(it.get("materialPrice")) for it in sov)
    exp = sum(num(it.get("quantity")) * num(it.get("expensePrice")) for it in sov)
    hours = cv.get("totalLaborHours") or hs.get("totalHours") or 0
    return {
        "n_items": len(sov), "material": mat, "expense": exp,
        "labor_hours": num(hours), "men": num(hs.get("men")),
        "sov": sov, "client": d.get("client", ""),
        "rates": d.get("hourlyRates") or {},
    }

# ---------- styles ----------
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
SUB_FONT = Font(italic=True, size=9, color="666666")
TOT_FONT = Font(bold=True, size=10)
TOT_FILL = PatternFill("solid", fgColor="D9E1F2")
ACT_FILL = PatternFill("solid", fgColor="FCE4D6")  # actuals-to-fill (QBO) = peach
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '#,##0;[Red](#,##0)'
PCT = '0.0%'
NUM0 = '#,##0'

wb = openpyxl.Workbook()

def style_header(ws, row, ncols, start=1):
    for c in range(start, start + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ===================================================================
# 1. READ ME
# ===================================================================
ws = wb.active; ws.title = "Read Me"
ws.sheet_view.showGridLines = False
ws["A1"] = "AMP Quality Energy Services — Job Profitability Model"; ws["A1"].font = TITLE_FONT
ws["A2"] = f"Bid/quote data pulled from ampOS (ampos.io) on {PULL_DATE}. Actuals to be sourced from QuickBooks (AMPQES)."
ws["A2"].font = SUB_FONT
notes = [
    ("", ""),
    ("PURPOSE", "Compare what AMP BID (ampOS) against what AMP ACTUALLY made (QuickBooks), per job — plus win/loss & pipeline analytics and estimate accuracy."),
    ("", ""),
    ("TABS", ""),
    ("  Pipeline & Win/Loss", "All 2,101 opportunities summarized by status, division, type, month. Win rate and $ at each stage. (Built entirely from ampOS.)"),
    ("  Bids (Opportunities)", "One row per quote/bid: amount, status, type, division, customer, win probability, whether an estimate exists, linked job #."),
    ("  Quote Line Items", "The detailed quote breakdown — one row per estimate line item (material, expense, labor men/hours, notes)."),
    ("  Won Jobs — Profitability", "One row per won job, joined to its bid. Bid value + estimated hours/material vs. ACTUAL revenue/cost (QBO, peach cells) → gross margin & variance."),
    ("  Estimate Accuracy", "Per job: estimated labor hours & cost vs. actual (QBO) — how good were the bids."),
    ("", ""),
    ("KEY DATA NOTES", ""),
    ("  Bid value", "= opportunities.quoted_amount (the NET-30 quoted price). The estimate JSON's stored grandTotal/calculatedValues are UNRELIABLE (store labor hours, not $) so are NOT used for $ — only line items and labor HOURS come from the estimate."),
    ("  Actuals", "Peach-shaded cells are placeholders to be filled from QuickBooks (AMPQES). jobs.quickbooks_project_id is currently EMPTY in ampOS, so QBO projects must be matched to jobs by name/customer."),
    ("  Division", "ampOS division values were inconsistent (e.g. 'Alabama' vs 'north_alabama') and have been normalized here."),
    ("  Labor rates", "ampOS estimates bill labor at ST $240 / OT $360 / DT $480 per hour (billing rate, not wage cost)."),
    ("  Coverage", "2,101 opportunities (Aug 2025–Jun 2026), 916 estimates, 145 active jobs, 2,269 customers."),
]
r = 3
for a, b in notes:
    ws.cell(row=r, column=1, value=a).font = Font(bold=bool(a) and not a.startswith("  "), size=10)
    ws.cell(row=r, column=2, value=b).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
set_widths(ws, [26, 120])
for rr in range(3, r):
    ws.row_dimensions[rr].height = 30

# ===================================================================
# 2. PIPELINE & WIN/LOSS
# ===================================================================
ws = wb.create_sheet("Pipeline & Win-Loss")
ws.sheet_view.showGridLines = False
ws["A1"] = "Pipeline & Win/Loss"; ws["A1"].font = TITLE_FONT
ws["A2"] = "All opportunities by stage. 'Won' = Awarded; 'Lost' = Lost; the rest are open pipeline."; ws["A2"].font = SUB_FONT

def summary_block(ws, top, title, keyfn, labelfn=None):
    agg = collections.defaultdict(lambda: [0, 0.0])  # key -> [count, quoted$]
    for o in opps:
        k = keyfn(o)
        agg[k][0] += 1
        agg[k][1] += num(o.get("quoted_amount"))
    ws.cell(row=top, column=1, value=title).font = Font(bold=True, size=11, color="1F3864")
    hdr = top + 1
    for i, h in enumerate(["Category", "# Bids", "Quoted $"], 1):
        ws.cell(row=hdr, column=i, value=h)
    style_header(ws, hdr, 3)
    rr = hdr + 1
    for k in sorted(agg, key=lambda x: -agg[x][1]):
        lab = labelfn(k) if labelfn else k
        ws.cell(row=rr, column=1, value=lab)
        ws.cell(row=rr, column=2, value=agg[k][0]).number_format = NUM0
        ws.cell(row=rr, column=3, value=round(agg[k][1])).number_format = MONEY
        rr += 1
    # total
    ws.cell(row=rr, column=1, value="TOTAL").font = TOT_FONT
    ws.cell(row=rr, column=2, value=f"=SUM(B{hdr+1}:B{rr-1})").number_format = NUM0
    ws.cell(row=rr, column=3, value=f"=SUM(C{hdr+1}:C{rr-1})").number_format = MONEY
    for c in range(1, 4):
        ws.cell(row=rr, column=c).font = TOT_FONT
        ws.cell(row=rr, column=c).fill = TOT_FILL
    return rr + 2

nxt = summary_block(ws, 4, "By Status / Stage", lambda o: o.get("status"), lambda k: STATUS_LABEL.get(k, k or "(none)"))
nxt = summary_block(ws, nxt, "By Division", lambda o: norm_div(o.get("amp_division")))
nxt = summary_block(ws, nxt, "By Opportunity Type", lambda o: o.get("opportunity_type"), lambda k: TYPE_LABEL.get(k, k or "(none)"))
nxt = summary_block(ws, nxt, "By Month Created", lambda o: (o.get("created_at") or "")[:7])

# Win-rate KPI block (col E)
won = sum(1 for o in opps if o.get("status") == "awarded")
lost = sum(1 for o in opps if o.get("status") == "lost")
won_val = sum(num(o.get("quoted_amount")) for o in opps if o.get("status") == "awarded")
lost_val = sum(num(o.get("quoted_amount")) for o in opps if o.get("status") == "lost")
kpis = [
    ("KPIs (decided bids only)", ""),
    ("Awarded (#)", won),
    ("Lost (#)", lost),
    ("Win rate (count)", (won / (won + lost)) if (won + lost) else 0),
    ("Awarded value $", round(won_val)),
    ("Lost value $", round(lost_val)),
    ("Win rate ($)", (won_val / (won_val + lost_val)) if (won_val + lost_val) else 0),
]
er = 4
for lab, val in kpis:
    ws.cell(row=er, column=5, value=lab)
    cell = ws.cell(row=er, column=6, value=val)
    if "rate" in lab:
        cell.number_format = PCT
    elif isinstance(val, (int, float)):
        cell.number_format = MONEY if "$" in lab else NUM0
    if lab.startswith("KPIs"):
        ws.cell(row=er, column=5).font = Font(bold=True, size=11, color="1F3864")
    er += 1
set_widths(ws, [22, 12, 16, 3, 24, 16])

# ===================================================================
# 3. BIDS (OPPORTUNITIES)
# ===================================================================
ws = wb.create_sheet("Bids (Opportunities)")
cols = ["Quote #", "Title", "Customer", "Division", "Status", "Type",
        "Quoted $", "Est. Man-Hrs", "Win Prob", "Has Estimate", "# Line Items",
        "Est. Material $", "Created", "Awarded", "Job #", "Sales Person"]
ws.append([""])
ws["A1"] = "Bids (Opportunities)"; ws["A1"].font = TITLE_FONT
ws.append(cols)
style_header(ws, 2, len(cols))
job_by_oppid = {j["opportunity_id"]: j for j in jobs if j.get("opportunity_id")}
for o in sorted(opps, key=lambda x: -num(x.get("quoted_amount"))):
    e = latest_est(o["id"])
    bd = est_breakdown(e) if e else None
    j = job_by_oppid.get(o["id"])
    ws.append([
        o.get("quote_number"), o.get("title"), cust_name(o.get("customer_id")),
        norm_div(o.get("amp_division")), STATUS_LABEL.get(o.get("status"), o.get("status")),
        TYPE_LABEL.get(o.get("opportunity_type"), o.get("opportunity_type")),
        round(num(o.get("quoted_amount"))), round(num(o.get("total_man_hours"))) or (round(bd["labor_hours"]) if bd else None),
        num(o.get("probability")) / 100 if o.get("probability") else 0,
        "Yes" if e else "No", bd["n_items"] if bd else None,
        round(bd["material"]) if bd else None,
        (o.get("created_at") or "")[:10], (o.get("awarded_date") or "")[:10] if o.get("awarded_date") else "",
        j.get("job_number") if j else "", o.get("sales_person"),
    ])
# formats
for row in range(3, ws.max_row + 1):
    ws.cell(row=row, column=7).number_format = MONEY
    ws.cell(row=row, column=8).number_format = NUM0
    ws.cell(row=row, column=9).number_format = PCT
    ws.cell(row=row, column=11).number_format = NUM0
    ws.cell(row=row, column=12).number_format = MONEY
ws.freeze_panes = "A3"
ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{ws.max_row}"
set_widths(ws, [9, 38, 26, 12, 16, 16, 13, 11, 9, 11, 11, 14, 12, 12, 8, 24])

# ===================================================================
# 4. QUOTE LINE ITEMS
# ===================================================================
ws = wb.create_sheet("Quote Line Items")
cols = ["Quote #", "Opportunity", "Customer", "Line Item", "Qty",
        "Material $/u", "Expense $/u", "Material Ext $", "Expense Ext $",
        "Labor Men", "Labor Hrs", "Notes"]
ws["A1"] = "Quote Line Items (Estimate Breakdown)"; ws["A1"].font = TITLE_FONT
ws.append([""]); ws.append(cols)
style_header(ws, 3, len(cols))
opp_by_id = {o["id"]: o for o in opps}
nrows = 0
for e in ests:
    bd = est_breakdown(e)
    if not bd or not bd["sov"]:
        continue
    o = opp_by_id.get(e["opportunity_id"]) or {}
    for it in bd["sov"]:
        qty = num(it.get("quantity")); mp = num(it.get("materialPrice")); ep = num(it.get("expensePrice"))
        ws.append([
            o.get("quote_number"), o.get("title"), cust_name(o.get("customer_id")),
            it.get("item"), qty, mp, ep, qty * mp, qty * ep,
            num(it.get("laborMen")), num(it.get("laborHours")), it.get("notes"),
        ])
        nrows += 1
for row in range(4, ws.max_row + 1):
    for c in (6, 7, 8, 9):
        ws.cell(row=row, column=c).number_format = MONEY
    for c in (5, 10, 11):
        ws.cell(row=row, column=c).number_format = NUM0
ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A3:{get_column_letter(len(cols))}{ws.max_row}"
set_widths(ws, [9, 34, 22, 40, 7, 12, 12, 13, 13, 9, 9, 30])

# ===================================================================
# 5. WON JOBS — PROFITABILITY
# ===================================================================
ws = wb.create_sheet("Won Jobs - Profitability")
cols = ["Job #", "Title", "Customer", "Division", "Job Status",
        "Bid $ (Quoted)", "Job Budget $", "Est. Man-Hrs",
        "Actual Revenue $", "Actual Cost $", "Actual Gross Margin $", "Gross Margin %",
        "Rev vs Bid Var $", "QBO Project (match)"]
ws["A1"] = "Won Jobs — Profitability (Bid vs Actual)"; ws["A1"].font = TITLE_FONT
ws.append([""])
ws.cell(row=2, column=1)
ws.append([""])
ws["A2"] = "Peach columns (I:J, N) = fill from QuickBooks (AMPQES). Margins/variances auto-calc."; ws["A2"].font = SUB_FONT
ws.append(cols)
HEADROW = 3
style_header(ws, HEADROW, len(cols))
first = HEADROW + 1
for j in sorted(jobs, key=lambda x: -(x.get("job_number_numeric") or 0)):
    o = opp_by_id.get(j.get("opportunity_id")) or {}
    e = latest_est(j.get("opportunity_id")) if j.get("opportunity_id") else None
    bd = est_breakdown(e) if e else None
    rr = ws.max_row + 1
    ws.append([
        j.get("job_number"), j.get("title"), cust_name(j.get("customer_id")),
        norm_div(j.get("amp_division") or o.get("amp_division")), j.get("status"),
        round(num(o.get("quoted_amount"))), round(num(j.get("budget"))),
        round(num(j.get("estimated_man_hours")) or num(o.get("total_man_hours")) or (bd["labor_hours"] if bd else 0)),
        None, None,  # actual rev, actual cost (QBO)
        f"=I{rr}-J{rr}", f"=IF(I{rr}=0,\"\",K{rr}/I{rr})",
        f"=IF(I{rr}=0,\"\",I{rr}-F{rr})",
        j.get("quickbooks_project_name") or "",
    ])
last = ws.max_row
# totals row
tr = last + 1
ws.cell(row=tr, column=1, value="TOTAL").font = TOT_FONT
for col in (6, 7, 9, 10, 11):
    ws.cell(row=tr, column=col, value=f"=SUM({get_column_letter(col)}{first}:{get_column_letter(col)}{last})")
ws.cell(row=tr, column=12, value=f"=IF(I{tr}=0,\"\",K{tr}/I{tr})")
# formats
for row in range(first, tr + 1):
    for c in (6, 7, 9, 10, 11, 13):
        ws.cell(row=row, column=c).number_format = MONEY
    ws.cell(row=row, column=8).number_format = NUM0
    ws.cell(row=row, column=12).number_format = PCT
    if row <= last:
        for c in (9, 10, 14):
            ws.cell(row=row, column=c).fill = ACT_FILL
for c in range(1, len(cols) + 1):
    ws.cell(row=tr, column=c).font = TOT_FONT
    ws.cell(row=tr, column=c).fill = TOT_FILL
ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A{HEADROW}:{get_column_letter(len(cols))}{last}"
set_widths(ws, [8, 36, 24, 12, 14, 14, 13, 11, 15, 14, 17, 12, 15, 22])

# ===================================================================
# 6. ESTIMATE ACCURACY
# ===================================================================
ws = wb.create_sheet("Estimate Accuracy")
cols = ["Job #", "Title", "Division", "Est. Man-Hrs", "Actual Man-Hrs",
        "Hrs Variance", "Hrs Var %", "Est. Material $", "Actual Material $", "Material Var $"]
ws["A1"] = "Estimate Accuracy (Estimated vs Actual)"; ws["A1"].font = TITLE_FONT
ws.append([""])
ws["A2"] = "Peach columns = fill from QuickBooks / field actuals. Variance = Actual − Estimate."; ws["A2"].font = SUB_FONT
ws.append(cols)
style_header(ws, 3, len(cols))
fr = 4
for j in sorted(jobs, key=lambda x: -(x.get("job_number_numeric") or 0)):
    o = opp_by_id.get(j.get("opportunity_id")) or {}
    e = latest_est(j.get("opportunity_id")) if j.get("opportunity_id") else None
    bd = est_breakdown(e) if e else None
    rr = ws.max_row + 1
    est_h = round(num(j.get("estimated_man_hours")) or num(o.get("total_man_hours")) or (bd["labor_hours"] if bd else 0))
    est_m = round(bd["material"]) if bd else 0
    ws.append([
        j.get("job_number"), j.get("title"), norm_div(j.get("amp_division") or o.get("amp_division")),
        est_h, None, f"=IF(E{rr}=\"\",\"\",E{rr}-D{rr})", f"=IF(OR(E{rr}=\"\",D{rr}=0),\"\",(E{rr}-D{rr})/D{rr})",
        est_m, None, f"=IF(I{rr}=\"\",\"\",I{rr}-H{rr})",
    ])
last = ws.max_row
for row in range(fr, last + 1):
    for c in (4, 5, 6):
        ws.cell(row=row, column=c).number_format = NUM0
    ws.cell(row=row, column=7).number_format = PCT
    for c in (8, 9, 10):
        ws.cell(row=row, column=c).number_format = MONEY
    for c in (5, 9):
        ws.cell(row=row, column=c).fill = ACT_FILL
ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A3:{get_column_letter(len(cols))}{last}"
set_widths(ws, [8, 38, 12, 12, 13, 12, 10, 14, 15, 14])

wb.save(OUT)
print("Saved:", OUT)
print("Tabs:", wb.sheetnames)
print(f"Bids rows: {len(opps)} | Line items: {nrows} | Jobs: {len(jobs)}")
